%pip install openpyxl
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill

def calculate_combi_differences(file1, file2, week):
    # Read Excel files into pandas DataFrames
    df1 = pd.read_excel(file1)
    df2 = pd.read_excel(file2)

    # Merge DataFrames on 'Combi'
    merged_df = pd.merge(df1, df2, on='Combi', suffixes=('_file1', '_file2'))

    # Extract data for the specified week from merged DataFrame
    week_col = str(week)
    week_cols_file1 = [col for col in merged_df.columns if str(col).startswith(week_col) and str(col).endswith('_file1')]
    week_cols_file2 = [col for col in merged_df.columns if str(col).startswith(week_col) and str(col).endswith('_file2')]

    # Ensure that only columns for the specified week are selected
    df_combi = merged_df[['Combi'] + week_cols_file1 + week_cols_file2].copy()

    # Calculate differences between the two DataFrames
    df_combi['Difference'] = df_combi[week_cols_file2[0]].astype(float) - df_combi[week_cols_file1[0]].astype(float)

    # Identify combis present in file 1 but not in file 2
    combis_only_in_file1 = df1[~df1['Combi'].isin(df2['Combi'])]

    # Identify combis present in file 2 but not in file 1
    combis_only_in_file2 = df2[~df2['Combi'].isin(df1['Combi'])]

    # Write information to separate sheets in the same Excel workbook
    with pd.ExcelWriter('combi_differences.xlsx', engine='openpyxl') as writer:
        df_combi.to_excel(writer, sheet_name='Combi Differences', index=False)
        combis_only_in_file1.to_excel(writer, sheet_name='Combi Only in File 1', index=False)
        combis_only_in_file2.to_excel(writer, sheet_name='Combi Only in File 2', index=False)

        # Access the workbook and active worksheet
        workbook = writer.book
        header_fill = PatternFill(start_color='92D050', end_color='92D050', fill_type='solid')
        for sheet_name in writer.sheets:
            worksheet = writer.sheets[sheet_name]

            # Apply green background to header cells
            for cell in worksheet['1:1']:
                cell.fill = header_fill

            # Set column width for all columns
            for col in worksheet.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = (max_length + 2) * 1.2
                worksheet.column_dimensions[column].width = adjusted_width

        # Apply color to difference column based on value
        if 'Combi Differences' in writer.sheets:
            diff_worksheet = writer.sheets['Combi Differences']
            for idx, row in df_combi.iterrows():
                diff_cell = diff_worksheet.cell(row=idx + 2, column=df_combi.columns.get_loc('Difference') + 1)
                if row['Difference'] > 0:
                    diff_cell.fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')  # Green
                elif row['Difference'] < 0:
                    diff_cell.fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')  # Red

    return df_combi

# Example usage
file1 = r'file_1.xlsx'
file2 = r'file_2.xlsx'
week = "06.2024"

combi_differences = calculate_combi_differences(file1, file2, week)

print("Resulting Combi Differences for Week 6.2024:")
print(combi_differences)  
print("Output written to combi_differences.xlsx")
